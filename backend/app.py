"""
========================================================
  ENERGYBAE - Solar Load Proposal Generator
  Built by: AI Intern

  How it works:
    1. User uploads an electricity bill (PDF or image)
    2. Groq Vision AI reads the bill and extracts key data
    3. Backend uses the bundled Excel template as a base
    4. Extracted data is filled into a COPY of that template
    5. User downloads the filled Excel proposal
    
  NOTE: The Excel template is bundled in the backend.
        Users do NOT need to upload it — ever.
========================================================
"""

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import Response, HTMLResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
import uvicorn
from typing import List
import json
import openpyxl
from io import BytesIO
import os
import base64
from datetime import datetime

# ── OpenAI-compatible client (used for Groq) ──
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

# ── PyMuPDF for converting PDFs to images ──
try:
    import fitz
    FITZ_AVAILABLE = True
except ImportError:
    FITZ_AVAILABLE = False


#  CONFIG — All backend constants in one place

GROQ_API_KEY    = os.environ.get("GROQ_API_KEY", "YOUR_GROQ_API_KEY_HERE")
GROQ_MODEL      = "meta-llama/llama-4-scout-17b-16e-instruct"
GROQ_BASE_URL   = "https://api.groq.com/openai/v1"

# Path to the bundled Excel template (sits next to app.py)
TEMPLATE_PATH   = os.path.join(os.path.dirname(__file__), "templates", "Energybae_Customer_Proposal.xlsx")


app = FastAPI(title="Energybae Solar Calculator API")

# Add CORS so the frontend can call this API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # For production, restrict this to the frontend domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

# ── Serve the frontend as static files ──
FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend")
FRONTEND_DIR = os.path.abspath(FRONTEND_DIR)

@app.get("/", response_class=HTMLResponse)
async def serve_index():
    index_path = os.path.join(FRONTEND_DIR, "index.html")
    with open(index_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())

# Serve CSS and JS at root level (matches HTML relative paths)
@app.get("/style.css")
async def serve_css():
    return FileResponse(os.path.join(FRONTEND_DIR, "style.css"), media_type="text/css")

@app.get("/script.js")
async def serve_js():
    return FileResponse(os.path.join(FRONTEND_DIR, "script.js"), media_type="application/javascript")

# Mount assets folder for images
app.mount("/assets", StaticFiles(directory=os.path.join(FRONTEND_DIR, "assets")), name="assets")
app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")


#  HELPER: Load the bundled Excel template from disk

def load_template_bytes():
    """
    Read the Excel template that ships with the app.
    Returns the raw bytes of the file.
    Raises FileNotFoundError with a clear message if missing.
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(
            f"Excel template not found at: {TEMPLATE_PATH}\n"
            "Make sure 'Energybae_Customer_Proposal.xlsx' is in the same folder as app.py."
        )
    with open(TEMPLATE_PATH, "rb") as f:
        return f.read()

#  HELPER: PDF → list of base64-encoded PNG images
#  Groq Vision accepts images, not raw PDFs

def pdf_to_base64_images(pdf_bytes):
    if not FITZ_AVAILABLE:
        raise RuntimeError("pymupdf not installed. Run: pip install pymupdf")
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    for page in doc:
        pix = page.get_pixmap(dpi=200)   # 200 DPI for crisp OCR
        b64 = base64.b64encode(pix.tobytes("png")).decode("utf-8")
        images.append(b64)
    doc.close()
    return images


#  STEP 1: Extract bill data using Groq Vision AI

def extract_bill_data(files):
    """
    Sends one or more electricity bill images/PDFs to Groq Vision
    in a SINGLE API call so the model can cross-reference all pages.

    Args:
        files: list of (file_bytes, file_name) tuples

    Returns:
        dict with the extracted fields
    """
    if not OPENAI_AVAILABLE:
        raise RuntimeError("openai library not installed. Run: pip install openai")

    client = OpenAI(api_key=GROQ_API_KEY, base_url=GROQ_BASE_URL)

    # Build image parts — all files go into ONE message
    image_parts = []
    for file_bytes, file_name in files:
        _, ext = os.path.splitext(file_name)
        ext = ext.lower()

        if ext == ".pdf":
            # Convert every PDF page to a separate image
            for b64 in pdf_to_base64_images(file_bytes):
                image_parts.append({
                    "type": "image_url",
                    "image_url": {"url": f"data:image/png;base64,{b64}", "detail": "high"}
                })
        else:
            mime = "image/jpeg" if ext in [".jpg", ".jpeg"] else "image/png"
            b64 = base64.b64encode(file_bytes).decode("utf-8")
            image_parts.append({
                "type": "image_url",
                "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "high"}
            })

    if not image_parts:
        raise ValueError("No valid images could be prepared from the uploaded files.")

    # Groq allows up to 5 images per request
    if len(image_parts) > 5:
        image_parts = image_parts[:5]

    prompt = """
You are an expert at reading Indian electricity bills (MSEDCL).
You are given one or more images/pages of the same electricity bill.
Use ALL images together to extract the most complete and accurate data.

First, THINK STEP-BY-STEP:
1. Locate the Consumer Information section (Name, Number, Load, Phase).
2. Locate the Fixed Charges (Sthir Akar) which is usually an integer.
3. Locate the 'Billed / Consumption details' table. Carefully read the rows one by one. Identify the Month and the Units consumed. Check if any rows are shifted.
4. After thinking, output the final JSON inside a ```json block.

Extract EXACTLY these fields:
- Consumer Name
- Consumer Number: MUST be EXACTLY 12 digits. Do not add any extra digits.
- Fixed Charges (number only, e.g. 130)
- Sanctioned Load (e.g. "3.30KW" or "1KW")
- Connection Type (e.g. "90/ LT I Res 1-Phase" or "Three Phase")
- Monthly consumption history: list of last 12 months with Month (YYYY-MM-DD), Units consumed, Bill Amount

Return ONLY a JSON object, no markdown, no explanation:
{
  "consumer_name": "...",
  "consumer_number": "...",
  "fixed_charges": 130,
  "sanctioned_load": "3.30KW",
  "connection_type": "90/ LT I Res 1-Phase",
  "monthly_data": [
    {"month": "2025-02-01", "units": 99,  "bill_amount": null},
    {"month": "2025-03-01", "units": 151, "bill_amount": null},
    ...
    {"month": "2026-01-01", "units": 25,  "bill_amount": 320.45}
  ]
}

bill_amount: only the LATEST month has it — put null for all others.
Ensure the JSON is wrapped in ```json ... ``` blocks.
"""

    response = client.chat.completions.create(
        model=GROQ_MODEL,
        messages=[{
            "role": "user",
            "content": [*image_parts, {"type": "text", "text": prompt}]
        }],
        temperature=0.1,
    )

    raw = response.choices[0].message.content.strip()
    
    import re
    
    # Strategy 1: Look for ```json ... ``` block
    match = re.search(r"```json\s*(.*?)\s*```", raw, re.DOTALL)
    if match:
        json_str = match.group(1).strip()
    else:
        # Strategy 2: Find first { and last }
        start = raw.find("{")
        end = raw.rfind("}")
        if start != -1 and end != -1:
            json_str = raw[start:end+1]
        else:
            json_str = "{}"
    
    # Try parsing; if "Extra data" error, use raw_decode to get just the first object
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        decoder = json.JSONDecoder()
        # Skip any leading whitespace
        idx = 0
        while idx < len(json_str) and json_str[idx] in ' \t\n\r':
            idx += 1
        obj, _ = decoder.raw_decode(json_str, idx)
        return obj


#  STEP 2: Fill the bundled Excel template with extracted data
#  Uses openpyxl so all existing formulas are preserved.
#  Only writes to data-input cells — never touches formula cells.

def fill_excel_template(bill_data):
    """
    Loads the bundled template from disk, fills it with
    the bill data, and returns the filled file as BytesIO.

    Cell mapping (from template inspection):
      D1  = Consumer Name
      D2  = Consumer Number
      D3  = Fixed Charges (Rs.)
      D4  = Sanctioned Load (KW)
      D5  = Connection Type
      C9:C20  = Month dates
      D9:D20  = Units consumed
      E9:E20  = Bill Amount (latest month only)

    Formula cells (NEVER touched):
      D22 = AVERAGE(D9:D21)
      D23 = required kW
      D24 = solar panels
      D25 = solar capacity
      D26 = number of panels
      D29 = total solar capacity
      D30 = number of solar panels
    """
    template_bytes = load_template_bytes()

    wb = openpyxl.load_workbook(filename=BytesIO(template_bytes))
    ws = wb.active

    # ── Consumer info ──
    ws["D1"] = bill_data.get("consumer_name", "")
    ws["D2"] = bill_data.get("consumer_number", "")
    try:
        ws["D3"] = float(bill_data.get("fixed_charges", 130) or 130)
    except:
        ws["D3"] = 130
    ws["D4"] = bill_data.get("sanctioned_load", "")
    ws["D5"] = bill_data.get("connection_type", "")

    # ── Monthly consumption (rows 9–20) ──
    monthly_data = bill_data.get("monthly_data", [])
    
    # Sort chronologically by month
    try:
        monthly_data.sort(key=lambda x: x.get("month", ""))
    except Exception:
        pass
        
    # Clear the old data from the right table (G, H, I, J) and left table
    for r in range(9, 21):
        for c in range(3, 11):
            if c != 6: # keep column F (unit cost) if any
                ws.cell(row=r, column=c).value = None
        # Put a 0 in column H to prevent AVERAGE(H9:H21) from throwing #DIV/0!
        ws.cell(row=r, column=8).value = 0
                
    for i, entry in enumerate(monthly_data):
        row = 9 + i
        if row > 20:
            break
        try:
            ws.cell(row=row, column=3).value = datetime.strptime(entry["month"], "%Y-%m-%d")
        except Exception:
            ws.cell(row=row, column=3).value = entry.get("month", "")
            
        try:
            u = float(entry.get("units", 0) or 0)
        except:
            u = 0
        ws.cell(row=row, column=4).value = u
        
        bill_amt = entry.get("bill_amount", None)
        if bill_amt is not None:
            try:
                b = float(bill_amt)
            except:
                b = 0
            ws.cell(row=row, column=5).value = b

    # Force Excel to recalculate formulas when opened
    # openpyxl doesn't evaluate formulas, but we can ensure the workbook is set to auto-calculate.
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    # ── Return the filled workbook as bytes ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.post("/api/generate-proposal")
async def generate_proposal(files: List[UploadFile] = File(...)):
    """
    API Endpoint to upload bill files and return the filled Excel file.
    """
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")
    
    # Read file contents
    files_input = []
    for file in files:
        contents = await file.read()
        files_input.append((contents, file.filename))
        
    try:
        # 1. Extract data
        extracted_data = extract_bill_data(files_input)
        
        # 2. Fill Excel
        filled_excel_bytesio = fill_excel_template(extracted_data)
        
        import re
        customer_name = str(extracted_data.get("consumer_name", "Customer"))
        customer_name = re.sub(r'[\\/*?:"<>|\n\r]', "", customer_name)
        customer_name = customer_name.strip().replace(" ", "_")
        if not customer_name:
            customer_name = "Customer"
            
        file_name = f"Energybae_Solar_Proposal_{customer_name}.xlsx"
        
        # Return the Excel file
        headers = {
            'Content-Disposition': f'attachment; filename="{file_name}"',
            'X-Extracted-Data': json.dumps(extracted_data).replace('\n', ' ').replace('\r', '') # Send extracted data back in header if frontend wants to show it
        }
        
        return Response(
            content=filled_excel_bytesio.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
    except Exception as e:
        import traceback
        traceback.print_exc()  # This will print the actual error to your Render logs!
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
