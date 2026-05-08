"""
ENERGYBAE - Solar Load Proposal Generator

How it works:
  1. User uploads an electricity bill (PDF or image)
  2. Groq Vision AI reads the bill and extracts key data
  3. Extracted data is filled into a COPY of the bundled Excel template
  4. User downloads the filled Excel proposal
"""

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import Response, HTMLResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
import uvicorn
from typing import List
import json
import openpyxl
from io import BytesIO
import os
import base64
from datetime import datetime

try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    import fitz
    FITZ_AVAILABLE = True
except ImportError:
    FITZ_AVAILABLE = False

GROQ_API_KEY    = os.environ.get("GROQ_API_KEY", "YOUR_GROQ_API_KEY_HERE")
GROQ_MODEL      = "meta-llama/llama-4-scout-17b-16e-instruct"
GROQ_BASE_URL   = "https://api.groq.com/openai/v1"
TEMPLATE_PATH   = os.path.join(os.path.dirname(__file__), "templates", "Energybae_Customer_Proposal.xlsx")

app = FastAPI(title="Energybae Solar Calculator API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

FRONTEND_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend"))

@app.get("/", response_class=HTMLResponse)
async def serve_index():
    index_path = os.path.join(FRONTEND_DIR, "index.html")
    with open(index_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())

@app.get("/style.css")
async def serve_css():
    return FileResponse(os.path.join(FRONTEND_DIR, "style.css"), media_type="text/css")

@app.get("/script.js")
async def serve_js():
    return FileResponse(os.path.join(FRONTEND_DIR, "script.js"), media_type="application/javascript")

app.mount("/assets", StaticFiles(directory=os.path.join(FRONTEND_DIR, "assets")), name="assets")
app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")


def load_template_bytes():
    """Read the bundled Excel template from disk."""
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Excel template not found at: {TEMPLATE_PATH}")
    with open(TEMPLATE_PATH, "rb") as f:
        return f.read()


def pdf_to_base64_images(pdf_bytes):
    """Convert each page of a PDF to a base64-encoded PNG image."""
    if not FITZ_AVAILABLE:
        raise RuntimeError("pymupdf not installed. Run: pip install pymupdf")
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    for page in doc:
        pix = page.get_pixmap(dpi=200)
        b64 = base64.b64encode(pix.tobytes("png")).decode("utf-8")
        images.append(b64)
    doc.close()
    return images


def extract_bill_data(files):
    """
    Send one or more bill images/PDFs to Groq Vision in a single API call.
    Returns a dict with extracted consumer and monthly consumption data.
    """
    if not OPENAI_AVAILABLE:
        raise RuntimeError("openai library not installed. Run: pip install openai")

    client = OpenAI(api_key=GROQ_API_KEY, base_url=GROQ_BASE_URL)

    image_parts = []
    for file_bytes, file_name in files:
        _, ext = os.path.splitext(file_name)
        ext = ext.lower()

        if ext == ".pdf":
            for b64 in pdf_to_base64_images(file_bytes):
                image_parts.append({
                    "type": "image_url",
                    "image_url": {"url": f"data:image/png;base64,{b64}", "detail": "high"}
                })
        else:
            mime = "image/jpeg" if ext in [".jpg", ".jpeg"] else "image/png"
            b64 = base64.b64encode(file_bytes).decode("utf-8")
            image_parts.append({
                "type": "image_url",
                "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "high"}
            })

    if not image_parts:
        raise ValueError("No valid images could be prepared from the uploaded files.")

    if len(image_parts) > 5:
        image_parts = image_parts[:5]

    prompt = """
You are an expert at reading Indian electricity bills (MSEDCL).
You are given one or more images/pages of the same electricity bill.
Use ALL images together to extract the most complete and accurate data.

First, THINK STEP-BY-STEP:
1. Locate the Consumer Information section (Name, Number, Load, Phase).
2. Locate the Fixed Charges (Sthir Akar) which is usually an integer.
3. Locate the 'Billed / Consumption details' table. Carefully read the rows one by one. Identify the Month and the Units consumed. Check if any rows are shifted.
4. After thinking, output the final JSON inside a ```json block.

Extract EXACTLY these fields:
- Consumer Name
- Consumer Number: MUST be EXACTLY 12 digits. Do not add any extra digits.
- Fixed Charges (number only, e.g. 130)
- Sanctioned Load (e.g. "3.30KW" or "1KW")
- Connection Type (e.g. "90/ LT I Res 1-Phase" or "Three Phase")
- Monthly consumption history: list of last 12 months with Month (YYYY-MM-DD), Units consumed, Bill Amount

Return ONLY a JSON object, no markdown, no explanation:
{
  "consumer_name": "...",
  "consumer_number": "...",
  "fixed_charges": 130,
  "sanctioned_load": "3.30KW",
  "connection_type": "90/ LT I Res 1-Phase",
  "monthly_data": [
    {"month": "2025-02-01", "units": 99,  "bill_amount": null},
    {"month": "2025-03-01", "units": 151, "bill_amount": null},
    ...
    {"month": "2026-01-01", "units": 25,  "bill_amount": 320.45}
  ]
}

bill_amount: only the LATEST month has it — put null for all others.
Ensure the JSON is wrapped in ```json ... ``` blocks.
"""

    response = client.chat.completions.create(
        model=GROQ_MODEL,
        messages=[{
            "role": "user",
            "content": [*image_parts, {"type": "text", "text": prompt}]
        }],
        temperature=0.1,
    )

    raw = response.choices[0].message.content.strip()
    
    import re
    
    match = re.search(r"```json\s*(.*?)\s*```", raw, re.DOTALL)
    if match:
        json_str = match.group(1).strip()
    else:
        start = raw.find("{")
        end = raw.rfind("}")
        if start != -1 and end != -1:
            json_str = raw[start:end+1]
        else:
            json_str = "{}"
    
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        decoder = json.JSONDecoder()
        idx = 0
        while idx < len(json_str) and json_str[idx] in ' \t\n\r':
            idx += 1
        obj, _ = decoder.raw_decode(json_str, idx)
        return obj


def fill_excel_template(bill_data):
    """
    Fill the bundled Excel template with extracted bill data.
    Only writes to data-input cells — never touches formula cells.

    Cell mapping:
      D1=Consumer Name, D2=Consumer Number, D3=Fixed Charges,
      D4=Sanctioned Load, D5=Connection Type
      C9:C20=Month dates, D9:D20=Units, E9:E20=Bill Amount
    """
    template_bytes = load_template_bytes()
    wb = openpyxl.load_workbook(filename=BytesIO(template_bytes))
    ws = wb.active

    ws["D1"] = bill_data.get("consumer_name", "")
    ws["D2"] = bill_data.get("consumer_number", "")
    ws["D3"] = bill_data.get("fixed_charges", 130)
    ws["D4"] = bill_data.get("sanctioned_load", "")
    ws["D5"] = bill_data.get("connection_type", "")

    monthly_data = bill_data.get("monthly_data", [])
    try:
        monthly_data.sort(key=lambda x: x.get("month", ""))
    except Exception:
        pass
        
    for i, entry in enumerate(monthly_data):
        row = 9 + i
        if row > 20:
            break
        try:
            ws.cell(row=row, column=3).value = datetime.strptime(entry["month"], "%Y-%m-%d")
        except Exception:
            ws.cell(row=row, column=3).value = entry.get("month", "")
        ws.cell(row=row, column=4).value = entry.get("units", None)
        bill_amt = entry.get("bill_amount", None)
        if bill_amt is not None:
            ws.cell(row=row, column=5).value = bill_amt

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.post("/api/generate-proposal")
async def generate_proposal(files: List[UploadFile] = File(...)):
    """Upload bill files and return the filled Excel proposal."""
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")
    
    files_input = []
    for file in files:
        contents = await file.read()
        files_input.append((contents, file.filename))
        
    try:
        extracted_data = extract_bill_data(files_input)
        filled_excel_bytesio = fill_excel_template(extracted_data)
        
        customer_name = extracted_data.get("consumer_name", "Customer").replace(" ", "_")
        file_name = f"Energybae_Solar_Proposal_{customer_name}.xlsx"
        
        headers = {
            'Content-Disposition': f'attachment; filename="{file_name}"',
            'X-Extracted-Data': json.dumps(extracted_data)
        }
        
        return Response(
            content=filled_excel_bytesio.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
