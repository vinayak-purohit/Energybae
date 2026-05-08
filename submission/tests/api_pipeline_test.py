"""
Full API pipeline test: uploads sample bill images to /api/generate-proposal
and verifies the response Excel has all formulas intact.
"""
import sys, json, os
sys.stdout.reconfigure(encoding='utf-8')
import requests
import openpyxl

API_URL = "http://127.0.0.1:8000/api/generate-proposal"
BILL_DIR = os.path.join(os.path.dirname(__file__), "..", "docs", "archive")
BILL1 = os.path.join(BILL_DIR, "Copy of WhatsApp Image 2026-02-12 at 13.48.47.jpeg")
BILL2 = os.path.join(BILL_DIR, "Copy of WhatsApp Image 2026-02-12 at 13.48.47 (1).jpeg")
OUTPUT = os.path.join(os.path.dirname(__file__), "api_test_output.xlsx")

print("=" * 60)
print("  ENERGYBAE API PIPELINE TEST")
print("=" * 60)

print("\n[1/5] Checking bill images exist...")
for path in [BILL1, BILL2]:
    exists = os.path.exists(path)
    size = os.path.getsize(path) if exists else 0
    print(f"  {'OK' if exists else 'MISSING'} {os.path.basename(path)} ({size} bytes)")

print("\n[2/5] Uploading bills to API...")
files = [
    ('files', ('bill1.jpeg', open(BILL1, 'rb'), 'image/jpeg')),
    ('files', ('bill2.jpeg', open(BILL2, 'rb'), 'image/jpeg'))
]
r = requests.post(API_URL, files=files, timeout=120)
print(f"  Status: {r.status_code}")
print(f"  Content-Type: {r.headers.get('content-type', '')}")
print(f"  Response size: {len(r.content)} bytes")

if r.status_code != 200:
    print(f"  ERROR: {r.text}")
    sys.exit(1)

print("\n[3/5] Checking extracted data...")
ext_header = r.headers.get('X-Extracted-Data', '')
if ext_header:
    data = json.loads(ext_header)
    print(f"  Consumer Name:   {data.get('consumer_name')}")
    print(f"  Consumer Number: {data.get('consumer_number')}")
    print(f"  Fixed Charges:   {data.get('fixed_charges')}")
    print(f"  Sanctioned Load: {data.get('sanctioned_load')}")
    print(f"  Connection Type: {data.get('connection_type')}")
    monthly = data.get('monthly_data', [])
    print(f"  Monthly entries: {len(monthly)}")
    for m in monthly:
        print(f"    {m.get('month')} - {m.get('units')} units - bill: {m.get('bill_amount')}")
else:
    print("  WARNING: No X-Extracted-Data header found")

cd = r.headers.get('Content-Disposition', '')
print(f"\n[4/5] Download header: {cd}")

print("\n[5/5] Verifying output Excel...")
with open(OUTPUT, 'wb') as f:
    f.write(r.content)
print(f"  Saved to {OUTPUT}")

wb = openpyxl.load_workbook(OUTPUT)
ws = wb.active

print("\n  Data Cells:")
print(f"    D1 (Name):       {ws['D1'].value}")
print(f"    D2 (Number):     {ws['D2'].value}")
print(f"    D3 (Fixed):      {ws['D3'].value}")
print(f"    D4 (Load):       {ws['D4'].value}")
print(f"    D5 (Type):       {ws['D5'].value}")
for row in range(9, 21):
    month = ws.cell(row=row, column=3).value
    units = ws.cell(row=row, column=4).value
    bill = ws.cell(row=row, column=5).value
    if month or units:
        print(f"    Row {row}: month={month}, units={units}, bill={bill}")

formula_cells = ['D22', 'D23', 'D24', 'D25', 'D26', 'D29', 'D30']
print("\n  Formula Verification:")
all_ok = True
for c in formula_cells:
    cell = ws[c]
    if cell.data_type == 'f':
        print(f"    OK   {c} = {cell.value}")
    else:
        print(f"    FAIL {c} = {cell.value} (NOT a formula!)")
        all_ok = False

print("\n" + "=" * 60)
if all_ok:
    print("  ALL TESTS PASSED - Pipeline working correctly!")
else:
    print("  SOME FORMULAS BROKEN!")
print("=" * 60)
