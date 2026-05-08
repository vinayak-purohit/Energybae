import openpyxl
import os
import sys

# Force utf-8 encoding for stdout
sys.stdout.reconfigure(encoding='utf-8')

def check_excel():
    f1 = "test_output_proposal.xlsx"
    f2 = "Energybae_Customer_Proposal.xlsx"
    f3 = "task/Copy of Pranay HOME E-Bill Analysis.xlsx"

    print("Loading workbooks...")
    wb1 = openpyxl.load_workbook(f1)
    wb2 = openpyxl.load_workbook(f2)
    wb3 = openpyxl.load_workbook(f3)

    ws1 = wb1.active
    ws2 = wb2.active
    ws3 = wb3.active

    print("\n--- Comparing Formulas between test_output_proposal and Energybae_Customer_Proposal ---")
    formula_cells = ["D22", "D23", "D24", "D25", "D26", "D29", "D30"]
    for cell_addr in formula_cells:
        c1 = ws1[cell_addr]
        c2 = ws2[cell_addr]
        c3 = ws3[cell_addr]

        v1 = c1.value if c1.data_type == 'f' else f"Not formula ({c1.value})"
        v2 = c2.value if c2.data_type == 'f' else f"Not formula ({c2.value})"
        v3 = c3.value if c3.data_type == 'f' else f"Not formula ({c3.value})"

        print(f"Cell {cell_addr}:")
        print(f"  test_output_proposal: {v1}")
        print(f"  Energybae_Customer_Proposal: {v2}")
        print(f"  Copy of Pranay: {v3}")
        if v1 == v2 and v1 == v3:
            print("  [PASS] Formulas match across all three.")
        else:
            print("  [FAIL] Formula mismatch!")

    print("\n--- Checking Data Values in Data Input Columns (test vs Copy of Pranay) ---")
    print("Checking rows 9 to 20 for Month (C), Units (D), Bill Amount (E)...")
    for row in range(9, 21):
        c_month1 = ws1.cell(row=row, column=3).value
        c_units1 = ws1.cell(row=row, column=4).value
        c_bill1 = ws1.cell(row=row, column=5).value

        c_month3 = ws3.cell(row=row, column=3).value
        c_units3 = ws3.cell(row=row, column=4).value
        c_bill3 = ws3.cell(row=row, column=5).value

        # format date if needed for comparison display
        def fmt(v):
            return str(v.date()) if hasattr(v, 'date') else str(v)

        m_match = (fmt(c_month1) == fmt(c_month3))
        u_match = (str(c_units1) == str(c_units3))
        b_match = (str(c_bill1) == str(c_bill3))

        print(f"Row {row}:")
        print(f"  test_output - Month: {fmt(c_month1)}, Units: {c_units1}, Bill: {c_bill1}")
        print(f"  Pranay      - Month: {fmt(c_month3)}, Units: {c_units3}, Bill: {c_bill3}")
        if m_match and u_match and b_match:
            print("  [PASS] Values match")
        else:
            print("  [FAIL] Values mismatch")

    print("\n--- Checking Other Customer Info ---")
    info_cells = ["D1", "D2", "D3", "D4", "D5"]
    for cell in info_cells:
        v1 = ws1[cell].value
        v3 = ws3[cell].value
        if str(v1) == str(v3):
            print(f"  [PASS] Cell {cell}: {v1} matches Pranay.")
        else:
            print(f"  [FAIL] Cell {cell}: test_output={v1}, Pranay={v3}")

if __name__ == '__main__':
    check_excel()
