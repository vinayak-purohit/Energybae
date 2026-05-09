"""
End-to-end test: Groq Vision extraction + Excel filling
Runs outside Streamlit to test the core pipeline directly.
"""
import sys, json, base64, os
sys.stdout.reconfigure(encoding='utf-8')

from openai import OpenAI
import openpyxl
from io import BytesIO
from datetime import datetime

API_KEY = os.environ.get("GROQ_API_KEY", "YOUR_GROQ_API_KEY_HERE")
BILL_IMAGES = [
    r"task\Copy of WhatsApp Image 2026-02-12 at 13.48.47.jpeg",
    r"task\Copy of WhatsApp Image 2026-02-12 at 13.48.47 (1).jpeg"
]
EXCEL_TEMPLATE = r"Energybae_Customer_Proposal.xlsx"
OUTPUT_FILE = r"test_output_proposal.xlsx"

print("=" * 60)
print("  ENERGYBAE E2E TEST (GROQ)")
print("=" * 60)

# ── STEP 1: Read the bill images ──
print("\n[1/4] Reading bill images...")
image_parts = []
for img_path in BILL_IMAGES:
    with open(img_path, "rb") as f:
        img_bytes = f.read()
    b64 = base64.b64encode(img_bytes).decode("utf-8")
    image_parts.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}", "detail": "high"}})
    print(f"  Loaded {img_path}: {len(img_bytes)} bytes")

# ── STEP 2: Call Groq Vision ──
print("\n[2/4] Calling Groq Vision API...")
client = OpenAI(api_key=API_KEY, base_url="https://api.groq.com/openai/v1")

prompt = """
You are an expert at reading Indian electricity bills (MSEDCL).
You have been provided with multiple pages/images of the same electricity bill. Cross-reference them to extract the most accurate data.

First, THINK STEP-BY-STEP:
1. Locate the Consumer Information section (Name, Number, Load, Phase).
2. Locate the Fixed Charges (Sthir Akar) which is usually an integer.
3. Locate the 'Billed / Consumption details' table. Carefully read the rows one by one. Identify the Month and the Units consumed. Check if any rows are shifted.
4. After thinking, output the final JSON inside a ```json block.

Extract EXACTLY these fields carefully:
- Consumer Name: (First, Middle, Last) Exactly as written.
- Consumer Number: MUST be EXACTLY 12 digits (e.g. 439320095567). Do not add any extra digits.
- Fixed Charges: (Sthir Akar) Usually between 100 to 500 Rs. Look closely for 'Fixed Charges' or 'Sthir Akar'. Extract only the number.
- Sanctioned Load: (e.g., "3.30KW" or "1KW")
- Connection Type: Look for Tariff/Phase. It usually says '1-Phase' or '90/ LT I Res 1-Phase'. Extract the exact phrase.
- Monthly consumption history: A table of the last 12 months with Month (as YYYY-MM-DD), Units, and Bill Amount. 
CRITICAL FOR TABLE: Align the rows perfectly. Trace the lines horizontally. Ensure units align with the correct month.

Use this exact JSON format in your final output:
{
  "consumer_name": "...",
  "consumer_number": "...",
  "fixed_charges": 130,
  "sanctioned_load": "3.30KW",
  "connection_type": "90/ LT I Res 1-Phase",
  "monthly_data": [
    {"month": "2025-02-01", "units": 99, "bill_amount": null},
    {"month": "2025-03-01", "units": 151, "bill_amount": null},
    ...
    {"month": "2026-01-01", "units": 25, "bill_amount": 320.45}
  ]
}

For bill_amount: only the LATEST month usually has it. Put null for others.
Ensure the JSON is wrapped in ```json ... ``` blocks.
"""

response = client.chat.completions.create(
    model="meta-llama/llama-4-scout-17b-16e-instruct",
    messages=[{
        "role": "user",
        "content": image_parts + [{"type": "text", "text": prompt}]
    }],
    temperature=0.1,
)

raw_text = response.choices[0].message.content.strip()

# Extract JSON block
import re
match = re.search(r"```json\n(.*?)\n```", raw_text, re.DOTALL)
if match:
    json_str = match.group(1).strip()
else:
    # Fallback: find first { and last }
    start = raw_text.find("{")
    end = raw_text.rfind("}")
    if start != -1 and end != -1:
        json_str = raw_text[start:end+1]
    else:
        json_str = "{}"

print(f"  Raw response length: {len(raw_text)} chars")

data = json.loads(json_str)
print("\n  ✅ Extracted data:")
print(f"    Consumer Name: {data.get('consumer_name')}")
print(f"    Consumer Number: {data.get('consumer_number')}")
print(f"    Fixed Charges: {data.get('fixed_charges')}")
print(f"    Sanctioned Load: {data.get('sanctioned_load')}")
print(f"    Connection Type: {data.get('connection_type')}")
print(f"    Monthly entries: {len(data.get('monthly_data', []))}")

# ── STEP 3: Fill Excel template ──
print("\n[3/4] Filling Excel template...")
with open(EXCEL_TEMPLATE, "rb") as f:
    template_bytes = f.read()

wb = openpyxl.load_workbook(filename=BytesIO(template_bytes))
ws = wb.active

ws["D1"] = data.get("consumer_name", "")
ws["D2"] = data.get("consumer_number", "")
ws["D3"] = data.get("fixed_charges", 130)
ws["D4"] = data.get("sanctioned_load", "")
ws["D5"] = data.get("connection_type", "")

monthly_data = data.get("monthly_data", [])

try:
    monthly_data.sort(key=lambda x: x.get("month", ""))
except Exception:
    pass

start_row = 9
for i, entry in enumerate(monthly_data):
    row = start_row + i
    if row > 20:
        break
    try:
        month_date = datetime.strptime(entry["month"], "%Y-%m-%d")
        ws.cell(row=row, column=3).value = month_date
    except Exception:
        ws.cell(row=row, column=3).value = entry.get("month", "")
    ws.cell(row=row, column=4).value = entry.get("units", None)
    bill_amt = entry.get("bill_amount", None)
    if bill_amt is not None:
        ws.cell(row=row, column=5).value = bill_amt

# ── STEP 4: Save and verify ──
wb.save(OUTPUT_FILE)
print(f"  ✅ Saved to {OUTPUT_FILE}")

# Verify formulas are intact
wb2 = openpyxl.load_workbook(OUTPUT_FILE)
ws2 = wb2.active
formula_cells = ["D22", "D23", "D24", "D25", "D26", "D29", "D30"]
print("\n[4/4] Verifying formulas preserved...")
all_ok = True
for cell_addr in formula_cells:
    cell = ws2[cell_addr]
    if cell.data_type == 'f':
        print(f"  ✅ {cell_addr} = {cell.value} (formula)")
    else:
        print(f"  ❌ {cell_addr} = {cell.value} (NOT a formula!)")
        all_ok = False

print("\n" + "=" * 60)
if all_ok:
    print("  ✅ ALL TESTS PASSED!")
else:
    print("  ⚠️ SOME FORMULAS WERE OVERWRITTEN!")
print("=" * 60)
